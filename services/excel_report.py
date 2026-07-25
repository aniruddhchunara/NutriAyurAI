import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from database.database import get_patients

os.makedirs("reports", exist_ok=True)

wb = Workbook()

ws = wb.active

ws.title = "Patients"


ws.append([
    "Name",
    "Age",
    "Weight",
    "Height",
    "Activity Factor",
    "BMI"
])

header_fill = PatternFill(
    fill_type="solid",
    start_color="1F4E78"
)
header_font = Font(
    bold=True,
    color="FFFFFF",
    size=12
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

for cell in ws[1]:

    cell.fill = header_fill

    cell.font = header_font

    cell.border = thin_border

    cell.alignment = center


patients  = get_patients()

for patient in patients:

    bmi = patient.weight / ((patient.height / 100) ** 2)

    ws.append([
    patient.name,
    patient.age,
    patient.weight,
    patient.height,
    patient.activity_factor,
    round(bmi, 2)
])

for row in ws.iter_rows(min_row=2):

    for cell in row:
        cell.border = thin_border
        cell.alignment = center

for column in ws.columns:

    max_length = 0

    column_letter = column[0].column_letter

    for cell in column:

        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:

            pass

    ws.column_dimensions[column_letter].width = max_length + 4


wb.save("reports/patient_report.xlsx")

print("Excel report created successfully!")


